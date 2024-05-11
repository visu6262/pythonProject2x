import openpyxl
wb=openpyxl.load_workbook(r"C:\Users\91939\Downloads\visudata.xlsx")
print(type(wb))
print(wb.sheetnames)
sh1=wb["Sheet1"]
print(sh1)
print(type(sh1))
print(sh1.cell(1,1).value,end=" ")
print(sh1.cell(1,2).value)
print(sh1.cell(2,1).value,end=" ")
print(sh1.cell(2,2).value)


