import openpyxl
wb=openpyxl.load_workbook(r"C:\Users\91939\Downloads\visu1.xlsx")
print(type(wb))
print(wb.sheetnames)
sh1=wb["Sheet1"]
print(sh1)
print(type(sh1))
print(sh1.cell(1,1).value,end=" ")
print(sh1.cell(1,2).value)
print(sh1.cell(2,1).value,end=" ")
print(sh1.cell(2,2).value)
row=sh1.max_row
print(row)
coloum=sh1.max_column
print(coloum)
print("------------")
# for a in range(2,row+1):
#     for b in range(1,coloum+1):
#         print(sh1.cell(a,b).value)
aaa=sh1.iter_cols(values_only=True)
one=[]
# two=[]
for a in aaa:

    # two.append(b.value)

    print(a)
